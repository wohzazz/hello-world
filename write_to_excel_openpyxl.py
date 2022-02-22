def write_to_excel_openpyxl(data, column_formats
                           ,output_file = 'data/excel_output_openpyxl.xlsx' ,table_title = None ,row_offset =2 , col_offset = 0, left_align_col = []
                           ):
    """
    Writes to excel using openpyxl library
    Last updated 23.02.2022
    
    argments
    
    (required arguments)
    
    data: DataFrame of data to output to excel
    column_formats: a list of formats with length equal to the number of columns in the table defined as data. If default formatting or no formatting preferred set to 'Normal'.
        for example, for a 6 column table, columns formats could be defined as ['Normal', 'Normal','Normal','Normal','Normal','Percent']
    
    
    (option arguments with defaults)
    output_file = path + file name of output file. Defaults to 'data/excel_output_openpyxl.xlsx'
    table_title: Table heading appearing in cell A1 with slighly larger navy font. Defaults to None
    row_offset: row to output table. 2 refers to row 3. Defaults to row 3.
    col_offset: column to start to output table. 0 refers to column 2. Defaults to column 2.
    left_align_col: list of columns to left aligned. should be a subset of the full list of column names. Defaults to an empty list []
    
    Formats for column_formats (more can be found in official documentation https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles):

            'Comma'
            ‘Comma [0]’
            'Percent'

    
    """

    # import openpyxl module, for styling see https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet from the active attribute
    sheet = wb.active

    # define colors to be used later
    pink = "00FF00FF"
    green = "00008000"
    black = "000000"
    navy = "000080"
    white = "FFFFFF"

    thin = Side(border_style="thin", color=navy)
    double = Side(border_style="double", color=navy)

    # write title (if exists)
    cell = sheet.cell(row =1, column = 1)
    cell.value = table_title
    cell.font = Font(name="Calibri", size=15, color=navy, bold =True)

    # write table headings  
    for i, col in enumerate(df.columns):
        cell = sheet.cell(row =1+row_offset, column = 1+i+col_offset)
        cell.value = col
        cell.font = Font(name="Calibri", size=11, color=white, bold =True)
        cell.fill = PatternFill(start_color=black, end_color=black,
                                            fill_type = "solid")
        cell.alignment= Alignment(horizontal='left',
                                          vertical='top')
    # write table contents
    for r, j in df.iterrows():
        for n in range(len(j)):
    #         print(j[n])
            cell = sheet.cell(row =1+r+row_offset+1, column = 1+n+col_offset)
            cell.value = j[n]
            cell.style = column_formats[n]
            cell.font = Font(name="Calibri", size=11, color=black)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # apply columns to not apply default right align (defined in left_align_col)
            skip_n_list = [] # values to skip
            for col in df.columns:
                try:
                    skip_n_list += [left_align_col.index(col)]
                except:
                    pass
    #         skip_n_list
            if n not in skip_n_list:
                cell.alignment= Alignment(horizontal='right',
                                          vertical='center')
 
    wb.save(output_file)
    print('Data outputted to "{}" ...'.format(output_file))

# sample data for illustration of the function ##############################################################################################
df = pd.DataFrame([['A', 'Z1', '23', 1, 1.21, 59.12, 0.66]
                   , ['B', 'Z2', '24', 2, 1.32, 38.15, 0.58]
    
], columns = ['CAT1', 'CAT2', 'CAT3', 'INT1', 'FLOAT1', 'DOLLAR1', 'PERC1'])
df.info()
# display(df)
fdf(df) #use reusable 'format DataFrame function'
############################################################################################################################################
write_to_excel_openpyxl(data = df, column_formats = ['Normal','Normal','Normal','Normal','Normal','Normal', 'Percent']
                           ,output_file = 'data/excel_output_illustration.xlsx' ,table_title = 'Sample excel write demo' ,row_offset =2 , col_offset = 0, left_align_col = ['CAT1', 'CAT2', 'CAT3']
                           )